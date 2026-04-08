import json
import base64
import asyncio
import logging
import subprocess
import tempfile
from pathlib import Path

import asyncpg
import httpx

from config import settings
from services.s3 import S3Service
from services.chunker import chunk_text, chunk_pages, store_chunks

logger = logging.getLogger(__name__)

MISTRAL_OCR_URL = "https://api.mistral.ai/v1/ocr"
MAX_RETRIES = 3
RETRY_BACKOFF = [2, 5, 10]

OFFICE_TYPES = {"pptx", "ppt", "docx", "doc"}
IMAGE_TYPES = {"png", "jpg", "jpeg", "webp", "gif"}
OCR_TYPES = {"pdf"} | OFFICE_TYPES | IMAGE_TYPES


class OCRService:
    def __init__(self, s3: S3Service, pool: asyncpg.Pool):
        self._s3 = s3
        self._pool = pool
        self._semaphore = asyncio.Semaphore(3)

    async def process_document(self, document_id: str, user_id: str):
        async with self._semaphore:
            await self._do_process(document_id, user_id)

    async def _check_global_limits(self, document_id: str):
        if not settings.GLOBAL_OCR_ENABLED:
            raise ValueError("OCR processing is temporarily disabled by the administrator.")

        total_pages = await self._pool.fetchval(
            "SELECT COALESCE(SUM(page_count), 0) FROM documents WHERE NOT archived"
        )
        if total_pages >= settings.GLOBAL_MAX_PAGES:
            raise ValueError(
                f"Platform page limit reached ({settings.GLOBAL_MAX_PAGES:,} pages). "
                "Please contact the administrator."
            )

    async def _do_process(self, document_id: str, user_id: str):
        try:
            await self._check_global_limits(document_id)
            await self._set_status(document_id, "processing")

            doc = await self._pool.fetchrow(
                "SELECT filename, file_type, knowledge_base_id::text as kb_id "
                "FROM documents WHERE id = $1 AND user_id = $2",
                document_id, user_id,
            )
            if not doc:
                logger.error("Document %s not found for user %s", document_id, user_id)
                return

            ext = doc["filename"].rsplit(".", 1)[-1].lower() if "." in doc["filename"] else doc["file_type"]
            kb_id = doc["kb_id"]
            s3_source_key = f"{user_id}/{document_id}/source.{ext}"

            if ext in OFFICE_TYPES:
                await self._convert_and_process(document_id, user_id, kb_id, s3_source_key, ext)
            elif ext in IMAGE_TYPES:
                await self._process_image(document_id, user_id, s3_source_key, ext)
            elif ext == "pdf":
                await self._process_pdf(document_id, user_id, kb_id, s3_source_key)
            elif ext in ("html", "htm"):
                await self._process_html(document_id, user_id, kb_id, s3_source_key)
            elif ext in ("xlsx", "xls", "csv"):
                await self._process_spreadsheet(document_id, user_id, kb_id, s3_source_key, ext)
            else:
                raise ValueError(f"Unsupported file type: {ext}")

        except Exception as e:
            logger.exception("Processing failed for document %s", document_id)
            try:
                await self._pool.execute(
                    "UPDATE documents SET status = 'failed', error_message = $2, updated_at = now() "
                    "WHERE id = $1",
                    document_id, str(e)[:500],
                )
            except Exception:
                logger.exception("Failed to update status to failed for %s", document_id)

    async def _process_pdf(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str):
        if settings.PDF_BACKEND == "mistral":
            if not settings.MISTRAL_API_KEY:
                raise ValueError("MISTRAL_API_KEY not configured — cannot process PDFs")
            presigned_url = await self._s3.generate_presigned_get(s3_source_key)
            ocr_result = await self._call_mistral_ocr(presigned_url, "document_url")
            await self._store_ocr_result(document_id, user_id, kb_id, ocr_result)
        else:
            await self._process_pdf_oxide(document_id, user_id, kb_id, s3_source_key)

    @staticmethod
    def _extract_pdf_oxide(pdf_path: str) -> list[tuple[int, str]]:
        """Run pdf-oxide extraction in a single thread (Rust objects are not Send)."""
        from pdf_oxide import PdfDocument
        doc = PdfDocument(pdf_path)
        pages = []
        for i in range(doc.page_count()):
            md = doc.to_markdown(i, detect_headings=True, include_images=False)
            pages.append((i + 1, md))
        return pages

    async def _process_pdf_oxide(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str):
        """Extract PDF text via pdf-oxide (free, local, no API calls)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = Path(tmpdir) / "source.pdf"
            await self._s3.download_to_file(s3_source_key, str(pdf_path))

            page_contents = await asyncio.to_thread(self._extract_pdf_oxide, str(pdf_path))
            num_pages = len(page_contents)

            user_limits = await self._pool.fetchrow(
                "SELECT page_limit FROM users WHERE id = $1", user_id,
            )
            page_limit = user_limits["page_limit"] if user_limits else settings.QUOTA_MAX_PAGES

            if num_pages > settings.QUOTA_MAX_PAGES_PER_DOC:
                raise ValueError(
                    f"Document has {num_pages} pages, maximum is {settings.QUOTA_MAX_PAGES_PER_DOC}."
                )

            current_pages = await self._pool.fetchval(
                "SELECT COALESCE(SUM(page_count), 0) FROM documents "
                "WHERE user_id = $1 AND id != $2",
                user_id, document_id,
            )
            if current_pages + num_pages > page_limit:
                raise ValueError(
                    f"Page quota exceeded: {current_pages} existing + {num_pages} new "
                    f"exceeds your limit of {page_limit} pages."
                )

        conn = await self._pool.acquire()
        try:
            await conn.execute("DELETE FROM document_pages WHERE document_id = $1", document_id)
            await conn.executemany(
                "INSERT INTO document_pages (document_id, page, content) "
                "VALUES ($1, $2, $3)",
                [(document_id, num, md) for num, md in page_contents],
            )
        finally:
            await self._pool.release(conn)

        full_content = "\n\n---\n\n".join(md for _, md in page_contents)
        chunks = chunk_pages(page_contents)
        await store_chunks(self._pool, document_id, user_id, kb_id, chunks)

        await self._pool.execute(
            "UPDATE documents SET status = 'ready', content = $2, page_count = $3, parser = 'pdf_oxide', updated_at = now() "
            "WHERE id = $1",
            document_id, full_content, num_pages,
        )
        logger.info("PDF (pdf-oxide): doc=%s pages=%d chunks=%d", document_id[:8], num_pages, len(chunks))

    async def _convert_to_pdf(self, document_id: str, user_id: str, s3_source_key: str, ext: str) -> str:
        """Convert office file to PDF via converter service or local LibreOffice. Returns S3 key of the PDF."""
        pdf_key = f"{user_id}/{document_id}/converted.pdf"

        if settings.CONVERTER_URL:
            source_url = await self._s3.generate_presigned_get(s3_source_key)
            result_url = await self._s3.generate_presigned_put(pdf_key)
            headers = {}
            if settings.CONVERTER_SECRET:
                headers["Authorization"] = f"Bearer {settings.CONVERTER_SECRET}"
            async with httpx.AsyncClient(timeout=httpx.Timeout(180.0, connect=10.0)) as client:
                resp = await client.post(
                    f"{settings.CONVERTER_URL}/convert",
                    json={"source_url": source_url, "result_url": result_url, "source_ext": ext},
                    headers=headers,
                )
                resp.raise_for_status()
        else:
            with tempfile.TemporaryDirectory() as tmpdir:
                source_path = Path(tmpdir) / f"source.{ext}"
                await self._s3.download_to_file(s3_source_key, str(source_path))
                result = await asyncio.to_thread(
                    subprocess.run,
                    ["libreoffice", "--headless", "--norestore", "--convert-to", "pdf", "--outdir", tmpdir, str(source_path)],
                    capture_output=True, timeout=120,
                )
                if result.returncode != 0:
                    raise RuntimeError(f"LibreOffice conversion failed: {result.stderr.decode()[:300]}")
                pdf_path = Path(tmpdir) / "source.pdf"
                if not pdf_path.exists():
                    raise RuntimeError("LibreOffice did not produce a PDF")
                await self._s3.upload_file(pdf_key, str(pdf_path), "application/pdf")

        return pdf_key

    async def _convert_and_process(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str, ext: str):
        """Convert office file to PDF, then process the PDF."""
        pdf_key = await self._convert_to_pdf(document_id, user_id, s3_source_key, ext)

        if settings.PDF_BACKEND == "mistral":
            if not settings.MISTRAL_API_KEY:
                raise ValueError("MISTRAL_API_KEY not configured")
            presigned_url = await self._s3.generate_presigned_get(pdf_key)
            ocr_result = await self._call_mistral_ocr(presigned_url, "document_url")
            await self._store_ocr_result(document_id, user_id, kb_id, ocr_result)
        else:
            await self._process_pdf_oxide(document_id, user_id, kb_id, pdf_key)

    async def _process_image(self, document_id: str, user_id: str, s3_source_key: str, ext: str):
        """Images are stored as-is. No OCR. The MCP read tool returns them natively."""
        await self._pool.execute(
            "UPDATE documents SET status = 'ready', page_count = 1, parser = 'native', updated_at = now() "
            "WHERE id = $1",
            document_id,
        )
        logger.info("Image stored: doc=%s", document_id[:8])

    async def _process_html(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str):
        """Parse HTML with webmd parser, store markdown + tagged HTML."""
        from html_parser import Parser

        html_bytes = await self._s3.download_bytes(s3_source_key)
        raw_html = html_bytes.decode("utf-8", errors="replace")

        parser = Parser(raw_html, content_only=True)
        result = parser.parse()

        await parser.embed_images()
        tagged_html = parser.html()

        await self._s3.upload_bytes(
            f"{user_id}/{document_id}/tagged.html",
            tagged_html.encode("utf-8"),
            "text/html",
        )

        markdown_content = result.content
        chunks = chunk_text(markdown_content)
        await store_chunks(self._pool, document_id, user_id, kb_id, chunks)

        await self._pool.execute(
            "UPDATE documents SET status = 'ready', content = $2, page_count = 1, parser = 'webmd', updated_at = now() "
            "WHERE id = $1",
            document_id, markdown_content,
        )
        logger.info("HTML processed: doc=%s chunks=%d", document_id[:8], len(chunks))

    async def _process_spreadsheet(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str, ext: str):
        """Download spreadsheet, store each sheet as a document_page."""
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / f"source.{ext}"
            await self._s3.download_to_file(s3_source_key, str(source_path))

            sheets = await asyncio.to_thread(self._parse_sheets, str(source_path), ext)

            conn = await self._pool.acquire()
            try:
                await conn.execute("DELETE FROM document_pages WHERE document_id = $1", document_id)
                content_parts = []
                for i, (name, md) in enumerate(sheets, 1):
                    content_parts.append(f"## {name}\n\n{md}")
                    await conn.execute(
                        "INSERT INTO document_pages (document_id, page, content, elements) "
                        "VALUES ($1, $2, $3, $4)",
                        document_id, i, md,
                        json.dumps({"sheet_name": name}),
                    )
            finally:
                await self._pool.release(conn)

            full_content = "\n\n---\n\n".join(content_parts)
            page_contents = [(i + 1, md) for i, (_, md) in enumerate(sheets)]
            chunks = chunk_pages(page_contents)
            await store_chunks(self._pool, document_id, user_id, kb_id, chunks)

            await self._pool.execute(
                "UPDATE documents SET status = 'ready', content = $2, page_count = $3, parser = 'openpyxl', updated_at = now() "
                "WHERE id = $1",
                document_id, full_content, len(sheets),
            )
            logger.info("Spreadsheet processed: doc=%s sheets=%d chunks=%d", document_id[:8], len(sheets), len(chunks))

    @staticmethod
    def _rows_to_markdown(rows: list[list[str]], max_rows: int = 100) -> str:
        if not rows:
            return "(empty)"
        header = "| " + " | ".join(rows[0]) + " |"
        sep = "| " + " | ".join("---" for _ in rows[0]) + " |"
        data = rows[1:max_rows + 1]
        body = "\n".join("| " + " | ".join(r) + " |" for r in data)
        truncated = f"\n\n*({len(rows) - 1 - max_rows} more rows truncated)*" if len(rows) - 1 > max_rows else ""
        return f"{header}\n{sep}\n{body}{truncated}"

    @staticmethod
    def _parse_sheets(path: str, ext: str) -> list[tuple[str, str]]:
        """Returns list of (sheet_name, markdown_table) tuples."""
        import csv
        if ext == "csv":
            with open(path, newline="", encoding="utf-8", errors="replace") as f:
                rows = [[c for c in row] for row in csv.reader(f)]
            return [("Sheet1", OCRService._rows_to_markdown(rows))]

        try:
            import openpyxl
        except ImportError:
            return [("Error", "(openpyxl not installed)")]

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [[str(cell) if cell is not None else "" for cell in row] for row in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            sheets.append((name, OCRService._rows_to_markdown(rows)))
        wb.close()
        return sheets

    async def _store_ocr_result(self, document_id: str, user_id: str, kb_id: str, ocr_result: dict):
        ocr_json_bytes = json.dumps(ocr_result).encode()
        await self._s3.upload_bytes(f"{user_id}/{document_id}/ocr.json", ocr_json_bytes, "application/json")

        pages = ocr_result.get("pages", [])

        if len(pages) > settings.QUOTA_MAX_PAGES_PER_DOC:
            raise ValueError(
                f"Document has {len(pages)} pages, maximum is {settings.QUOTA_MAX_PAGES_PER_DOC}."
            )

        user_limits = await self._pool.fetchrow(
            "SELECT page_limit, storage_limit_bytes FROM users WHERE id = $1",
            user_id,
        )
        page_limit = user_limits["page_limit"] if user_limits else settings.QUOTA_MAX_PAGES

        current_pages = await self._pool.fetchval(
            "SELECT COALESCE(SUM(page_count), 0) FROM documents "
            "WHERE user_id = $1 AND id != $2",
            user_id, document_id,
        )
        if current_pages + len(pages) > page_limit:
            raise ValueError(
                f"Page quota exceeded: {current_pages} existing + {len(pages)} new "
                f"exceeds your limit of {page_limit} pages."
            )

        for page in pages:
            for img in page.get("images", []):
                img_id = img.get("id")
                img_b64 = img.get("image_base64")
                if not img_id or not img_b64:
                    continue
                if img_b64.startswith("data:"):
                    img_b64 = img_b64.split(",", 1)[1]
                img_bytes = base64.b64decode(img_b64)
                await self._s3.upload_bytes(
                    f"{user_id}/{document_id}/images/{img_id}",
                    img_bytes,
                    "image/jpeg",
                )

        content_parts = []
        conn = await self._pool.acquire()
        try:
            await conn.execute("DELETE FROM document_pages WHERE document_id = $1", document_id)
            for page in pages:
                page_index = page.get("index", 0) + 1
                page_md = page.get("markdown", "")
                content_parts.append(page_md)

                elements = {}
                if page.get("images"):
                    elements["images"] = [
                        {k: v for k, v in img.items() if k != "image_base64"}
                        for img in page["images"]
                    ]
                if page.get("dimensions"):
                    elements["dimensions"] = page["dimensions"]
                if page.get("tables"):
                    elements["tables"] = page["tables"]

                await conn.execute(
                    "INSERT INTO document_pages (document_id, page, content, elements) "
                    "VALUES ($1, $2, $3, $4)",
                    document_id, page_index, page_md,
                    json.dumps(elements) if elements else None,
                )
        finally:
            await self._pool.release(conn)

        full_content = "\n\n---\n\n".join(content_parts)
        page_count = len(pages)

        page_contents = [(page.get("index", 0) + 1, page.get("markdown", "")) for page in pages]
        chunks = chunk_pages(page_contents)
        await store_chunks(self._pool, document_id, user_id, kb_id, chunks)

        await self._pool.execute(
            "UPDATE documents SET status = 'ready', content = $2, page_count = $3, parser = 'mistral', updated_at = now() "
            "WHERE id = $1",
            document_id, full_content, page_count,
        )
        logger.info("OCR complete: doc=%s pages=%d chunks=%d", document_id[:8], page_count, len(chunks))

    async def _call_mistral_ocr(self, url: str, url_type: str = "document_url") -> dict:
        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                async with httpx.AsyncClient(timeout=httpx.Timeout(300.0, connect=10.0)) as client:
                    resp = await client.post(
                        MISTRAL_OCR_URL,
                        headers={
                            "Authorization": f"Bearer {settings.MISTRAL_API_KEY}",
                            "Content-Type": "application/json",
                        },
                        json={
                            "model": "mistral-ocr-latest",
                            "document": {
                                "type": url_type,
                                url_type: url,
                            },
                            "include_image_base64": True,
                            "table_format": "markdown",
                        },
                    )
                    resp.raise_for_status()
                    return resp.json()
            except (httpx.HTTPStatusError, httpx.TimeoutException) as e:
                last_error = e
                if attempt < MAX_RETRIES - 1:
                    wait = RETRY_BACKOFF[attempt]
                    logger.warning("Mistral OCR attempt %d failed: %s, retrying in %ds", attempt + 1, e, wait)
                    await asyncio.sleep(wait)
        raise last_error or RuntimeError("Mistral OCR failed after retries")

    async def _set_status(self, document_id: str, status: str):
        await self._pool.execute(
            "UPDATE documents SET status = $2, updated_at = now() WHERE id = $1",
            document_id, status,
        )
